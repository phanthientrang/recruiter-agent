import asyncio
import contextlib
import io
import json
import os
import re
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Annotated, Optional, TypedDict

from dotenv import load_dotenv
from langchain_core.messages import SystemMessage, ToolMessage
from langchain_core.tools import tool
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, START
from langgraph.graph.message import add_messages
from langgraph.prebuilt import ToolNode, tools_condition
from openpyxl import Workbook
from pydantic import BaseModel

from greennode_agentbase import GreenNodeAgentBaseApp, RequestContext, PingStatus
from starlette.middleware import Middleware
from starlette.middleware.cors import CORSMiddleware
from starlette.requests import Request
from starlette.responses import JSONResponse, Response
from starlette.routing import Route

load_dotenv()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TA_INCHARGE = os.environ.get("TA_INCHARGE") or os.environ.get("USERNAME", "Unknown")
JOB_SUB_FOLDERS = ["LinkedIn", "VNG Careers", "Referral", "TA Search", "Others"]
_ALLOWED_EXT = {".pdf", ".docx"}

DB_COLUMNS = [
    "No", "Request code", "Candidate name", "Processed Team", "Processed Position",
    "Entry date", "Source", "Referrer", "Latest company", "Latest position",
    "Email", "Phone", "Stage", "Status", "Note", "Reason for failure/withdrawal",
    "Last drawn salary", "Expected salary (Monthly Gross)", "TA Incharge", "Profile",
]
SYNC_COLUMNS = [c for c in DB_COLUMNS if c not in ("No", "Profile")]

# Schedule: TA_INCHARGE value (case-insensitive) -> (hour, minute)
SYNC_SCHEDULE: dict[str, tuple[int, int]] = {
    "trangptt12": (7, 30),
    "hautt2":     (8,  0),
    "huyenplt":   (8, 30),
    "nhihm":      (9,  0),
}

# ---------------------------------------------------------------------------
# In-memory store with local write-through (survives within one running
# container/process; reset on redeploy or restart — there is no generic
# AgentBase persistent-storage service, see project notes).
# ---------------------------------------------------------------------------
_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(_DATA_DIR, exist_ok=True)

PERSONAL_DB_PATH = os.path.join(_DATA_DIR, "personal_db.json")
TEAM_DB_PATH     = os.path.join(_DATA_DIR, "team_db.json")
JOB_FOLDERS_PATH = os.path.join(_DATA_DIR, "job_folders.json")
ALERTS_PATH      = os.path.join(_DATA_DIR, "alerts.json")
SYNC_STATE_PATH  = os.path.join(_DATA_DIR, "sync_state.json")

_store_lock = threading.Lock()


def _load_json(path: str, default):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def _save_json(path: str, data) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)


personal_db: list[dict] = _load_json(PERSONAL_DB_PATH, [])
team_db: list[dict]     = _load_json(TEAM_DB_PATH, [])
job_folders: list[dict] = _load_json(JOB_FOLDERS_PATH, [])
_alerts: list[dict]     = _load_json(ALERTS_PATH, [])
_sync_state: dict       = _load_json(SYNC_STATE_PATH, {"last_sync": None, "rows": {}})


def _persist_personal() -> None:
    with _store_lock:
        _save_json(PERSONAL_DB_PATH, personal_db)


def _persist_team() -> None:
    with _store_lock:
        _save_json(TEAM_DB_PATH, team_db)


def _persist_folders() -> None:
    with _store_lock:
        _save_json(JOB_FOLDERS_PATH, job_folders)


def _persist_alerts() -> None:
    with _store_lock:
        _save_json(ALERTS_PATH, _alerts)


def _persist_sync_state() -> None:
    with _store_lock:
        _save_json(SYNC_STATE_PATH, _sync_state)


def _push_alert(alert: dict) -> None:
    _alerts.append(alert)
    _persist_alerts()


def _remove_alert(alert: dict) -> None:
    _alerts.remove(alert)
    _persist_alerts()


def _clear_alerts() -> None:
    _alerts.clear()
    _persist_alerts()


def _next_no(db: list[dict]) -> int:
    return len(db) + 1


def _public_row(row: dict) -> dict:
    return {col: row.get(col) for col in DB_COLUMNS}


def _split_folder_name(name: str) -> dict | None:
    """Expects: '[Dept] - [Position] - [JobCode]'."""
    segments = name.split(" - ", 2)
    if len(segments) < 3:
        return None
    return {
        "processed_team":     segments[0].strip(),
        "processed_position":  segments[1].strip(),
        "request_code":        segments[2].strip(),
    }


# ---------------------------------------------------------------------------
# Skill 2 — CV text extraction (from uploaded bytes, no disk access)
# ---------------------------------------------------------------------------
def _extract_pdf_text_bytes(data: bytes) -> str:
    import pypdf
    reader = pypdf.PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx_text_bytes(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_cv_text_bytes(filename: str, data: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return _extract_pdf_text_bytes(data)
    if ext == ".docx":
        return _extract_docx_text_bytes(data)
    raise ValueError(f"Unsupported file type '{ext}'. Only PDF and DOCX are supported.")


# ---------------------------------------------------------------------------
# Skill 2 — LLM CV field extraction
# ---------------------------------------------------------------------------
class _CVFields(BaseModel):
    candidate_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    latest_company: Optional[str] = None
    latest_position: Optional[str] = None


def _parse_cv_with_llm(text: str) -> _CVFields:
    structured_llm = llm.with_structured_output(_CVFields)
    return structured_llm.invoke(
        "Extract the following fields from the CV below. "
        "Return null for any field that cannot be clearly determined — do NOT guess.\n\n"
        f"CV:\n{text[:6000]}"
    )


# ---------------------------------------------------------------------------
# Skill 2 — Personal DB operations (in-memory)
# ---------------------------------------------------------------------------
def _check_duplicate(email: str, request_code: str) -> dict | None:
    email_l, rc = email.strip().lower(), request_code.strip()
    for row in personal_db:
        if (str(row.get("Email") or "").strip().lower() == email_l
                and str(row.get("Request code") or "").strip() == rc):
            return row
    return None


def _check_cross_role_duplicate(email: str, exclude_request_code: str) -> list[dict]:
    email_l = email.strip().lower()
    exc_rc = exclude_request_code.strip()
    ta_l = (TA_INCHARGE or "").strip().lower()
    return [
        row for row in personal_db
        if str(row.get("Email") or "").strip().lower() == email_l
        and str(row.get("Request code") or "").strip() != exc_rc
        and str(row.get("TA Incharge") or "").strip().lower() == ta_l
    ]


def _add_personal_row(row_data: dict) -> dict:
    row = {col: row_data.get(col) for col in DB_COLUMNS}
    row["No"] = _next_no(personal_db)
    row["_id"] = uuid.uuid4().hex
    personal_db.append(row)
    _persist_personal()
    return row


def _overwrite_personal_row(row_id: str, row_data: dict) -> bool:
    for row in personal_db:
        if row.get("_id") == row_id:
            for col in DB_COLUMNS:
                if col != "No" and col in row_data:
                    row[col] = row_data[col]
            _persist_personal()
            return True
    return False


def _build_row_data(cv: _CVFields, folder_name: str, source: str, referrer: str | None, filename: str) -> dict:
    parts = _split_folder_name(folder_name) or {}
    return {
        "Request code":       parts.get("request_code"),
        "Candidate name":     cv.candidate_name,
        "Processed Team":     parts.get("processed_team"),
        "Processed Position": parts.get("processed_position"),
        "Entry date":         datetime.now().strftime("%Y-%m-%d"),
        "Source":             source,
        "Referrer":           referrer if source == "Referral" else None,
        "Latest company":     cv.latest_company,
        "Latest position":    cv.latest_position,
        "Email":               cv.email,
        "Phone":               cv.phone,
        "TA Incharge":         TA_INCHARGE,
        "Profile":             f"{filename} - uploaded {datetime.now().strftime('%Y-%m-%d')}",
    }


# ---------------------------------------------------------------------------
# Skill 2 — Core CV processing pipeline (operates on uploaded bytes)
# ---------------------------------------------------------------------------
def _process_cv(filename: str, data: bytes, folder_name: str, subfolder: str,
                 referrer: str | None = None) -> dict:
    result: dict = {"file": filename, "status": None, "messages": []}

    try:
        text = _extract_cv_text_bytes(filename, data)
    except Exception as e:
        result.update(status="error", messages=[f"Cannot extract text: {e}"])
        _push_alert({**result, "timestamp": datetime.now().isoformat()})
        return result

    if not text.strip():
        result.update(status="error", messages=["CV appears empty or image-only — could not extract text."])
        _push_alert({**result, "timestamp": datetime.now().isoformat()})
        return result

    folder = _split_folder_name(folder_name)
    if not folder:
        result.update(status="error", messages=["Could not parse job folder name. Expected: '[Dept] - [Position] - [JobCode]'"])
        _push_alert({**result, "timestamp": datetime.now().isoformat()})
        return result

    try:
        cv = _parse_cv_with_llm(text)
    except Exception as e:
        result.update(status="error", messages=[f"LLM extraction failed: {e}"])
        _push_alert({**result, "timestamp": datetime.now().isoformat()})
        return result

    missing = [label for label, val in [
        ("Candidate name", cv.candidate_name), ("Email", cv.email), ("Phone", cv.phone),
        ("Latest company", cv.latest_company), ("Latest position", cv.latest_position),
    ] if not val]
    if missing:
        result["messages"].append(f"Warning - Missing fields (left blank): {', '.join(missing)}")

    new_job = " - ".join(filter(None, [folder.get("processed_team"), folder.get("processed_position"), folder.get("request_code")]))

    if cv.email:
        dup = _check_duplicate(cv.email, folder["request_code"])
        if dup:
            existing_job = " - ".join(filter(None, [dup.get("Processed Team"), dup.get("Processed Position"), folder["request_code"]]))
            result.update(status="duplicate", messages=result["messages"] + [
                f"DUPLICATE: {cv.candidate_name or 'Unknown'} ({cv.email}) already exists for {folder['request_code']}. "
                "Use resolve_duplicate to keep, overwrite, or add as new."
            ], email=cv.email, candidate_name=cv.candidate_name, request_code=folder["request_code"],
               new_job=new_job, new_source=subfolder, existing_job=existing_job)
            _push_alert({**result, "timestamp": datetime.now().isoformat(),
                "cv_fields": cv.model_dump(), "folder_name": folder_name, "subfolder": subfolder,
                "referrer": referrer, "filename": filename, "duplicate_id": dup["_id"]})
            return result

        cross_role = _check_cross_role_duplicate(cv.email, folder["request_code"])
        if cross_role:
            codes = ", ".join(r["Request code"] for r in cross_role)
            first = cross_role[0]
            existing_job = " - ".join(filter(None, [first.get("Processed Team"), first.get("Processed Position"), first.get("Request code")]))
            result.update(status="cross_role_duplicate", messages=result["messages"] + [
                f"CROSS-ROLE: {cv.candidate_name or 'This candidate'} ({cv.email}) is already "
                f"in your pipeline for another role ({codes}). "
                "Use resolve_duplicate to keep or add as new."
            ], email=cv.email, candidate_name=cv.candidate_name, request_code=folder["request_code"],
               new_job=new_job, new_source=subfolder, existing_job=existing_job)
            _push_alert({**result, "timestamp": datetime.now().isoformat(),
                "cv_fields": cv.model_dump(), "folder_name": folder_name, "subfolder": subfolder,
                "referrer": referrer, "filename": filename,
                "existing_roles": [r["Request code"] for r in cross_role]})
            return result

    row_data = _build_row_data(cv, folder_name, subfolder, referrer, filename)
    if not cv.email:
        row_data["Note"] = "Email missing - please fill in manually"
    new_row = _add_personal_row(row_data)
    result.update(status="success")
    result["messages"].append(f"Added {cv.candidate_name or 'candidate'} to database (No. {new_row['No']}).")
    _push_alert({**result, "timestamp": datetime.now().isoformat()})
    return result


# ---------------------------------------------------------------------------
# Skill 3 — Sync helpers (personal DB -> team DB, both in-memory)
# ---------------------------------------------------------------------------
def _row_key(row: dict) -> str:
    email = str(row.get("Email") or "").strip().lower()
    code = str(row.get("Request code") or "").strip()
    return f"{email}|{code}"


def _row_changed_vs_snapshot(current: dict, snapshot: dict) -> bool:
    return any(str(current.get(c) or "").strip() != str(snapshot.get(c) or "").strip() for c in SYNC_COLUMNS)


def _add_team_row(row_data: dict) -> dict:
    row = {col: row_data.get(col) for col in SYNC_COLUMNS}
    row["No"] = _next_no(team_db)
    row["Profile"] = None
    row["_id"] = uuid.uuid4().hex
    team_db.append(row)
    _persist_team()
    return row


def _update_team_row(row: dict, changed_fields: dict) -> None:
    for field, vals in changed_fields.items():
        row[field] = vals["mine"]
    _persist_team()


def _get_changed_fields(my_row: dict, team_row: dict) -> dict:
    return {
        col: {"mine": str(my_row.get(col) or "").strip(), "theirs": str(team_row.get(col) or "").strip()}
        for col in SYNC_COLUMNS
        if str(my_row.get(col) or "").strip() != str(team_row.get(col) or "").strip()
    }


def _run_sync() -> dict:
    today = datetime.now().strftime("%Y-%m-%d")
    result = {"date": today, "timestamp": datetime.now().isoformat(),
              "added": 0, "updated": 0, "skipped": 0, "conflicts": 0, "errors": []}

    team_lookup: dict[str, dict] = {_row_key(r): r for r in team_db if _row_key(r) not in ("|", "")}
    snapshot = _sync_state.get("rows", {})

    email_lookup: dict[str, list[dict]] = {}
    for row in team_db:
        em = str(row.get("Email") or "").strip().lower()
        if em:
            email_lookup.setdefault(em, []).append({
                "request_code": str(row.get("Request code") or "").strip(),
                "team_ta": str(row.get("TA Incharge") or "").strip().lower(),
                "team_ta_display": str(row.get("TA Incharge") or ""),
            })

    to_sync = []
    for row in personal_db:
        entry_date = str(row.get("Entry date") or "").strip()[:10]
        key = _row_key(row)
        if not key or key == "|":
            continue
        if entry_date == today or (key in snapshot and _row_changed_vs_snapshot(row, snapshot[key])):
            to_sync.append(row)

    def _snapshot_now() -> dict:
        return {_row_key(r): {k: str(v) if v is not None else None for k, v in r.items()}
                for r in personal_db if _row_key(r) not in ("|", "")}

    if not to_sync:
        _sync_state["last_sync"] = datetime.now().isoformat()
        _sync_state["rows"] = _snapshot_now()
        _persist_sync_state()
        summary = f"Daily sync {today}: No changes to sync today."
        _push_alert({"status": "sync_complete", "messages": [summary],
                     "timestamp": result["timestamp"], "details": result})
        return result

    for my_row in to_sync:
        key = _row_key(my_row)
        email = str(my_row.get("Email") or "").strip().lower()
        request_code = str(my_row.get("Request code") or "").strip()
        candidate_name = str(my_row.get("Candidate name") or "")
        serialized_row = {k: str(v) if v is not None else None for k, v in my_row.items()}
        try:
            if key not in team_lookup:
                cross_ta = [
                    e for e in email_lookup.get(email, [])
                    if e["request_code"] != request_code
                    and e["team_ta"] != (TA_INCHARGE or "").lower().strip()
                ] if email else []
                if cross_ta:
                    codes = ", ".join(e["request_code"] for e in cross_ta)
                    tas = ", ".join(sorted(set(e["team_ta_display"] for e in cross_ta)))
                    _push_alert({
                        "status": "sync_cross_ta_pending", "timestamp": datetime.now().isoformat(),
                        "email": email, "request_code": request_code, "candidate_name": candidate_name,
                        "team_ta": tas, "row_data": serialized_row,
                        "messages": [f"Cross-TA: {candidate_name} ({email}) applying across team - "
                                     f"{tas} already has this candidate for {codes}. "
                                     "Use resolve_sync_conflict to sync or skip."],
                    })
                    result["conflicts"] += 1
                else:
                    new_team_row = _add_team_row(my_row)
                    team_lookup[key] = new_team_row
                    result["added"] += 1
            else:
                team_row = team_lookup[key]
                changed = _get_changed_fields(my_row, team_row)
                if not changed:
                    result["skipped"] += 1
                    continue
                team_ta = str(team_row.get("TA Incharge") or "").strip().lower()
                my_ta = (TA_INCHARGE or "").lower().strip()
                if team_ta != my_ta:
                    _push_alert({
                        "status": "sync_conflict_pending", "timestamp": datetime.now().isoformat(),
                        "email": email, "request_code": request_code, "candidate_name": candidate_name,
                        "team_ta": str(team_row.get("TA Incharge") or ""), "row_data": serialized_row,
                        "messages": [f"Sync conflict: {candidate_name} ({email}) for {request_code} - "
                                     f"{team_row.get('TA Incharge')} already has this candidate in team DB. "
                                     "Use resolve_sync_conflict."],
                    })
                    result["conflicts"] += 1
                else:
                    _update_team_row(team_row, changed)
                    result["updated"] += 1
        except Exception as e:
            result["errors"].append(f"{candidate_name} ({email}): {e}")

    _sync_state["last_sync"] = datetime.now().isoformat()
    _sync_state["rows"] = _snapshot_now()
    _persist_sync_state()

    summary = (f"Daily sync {today}: {result['added']} added, {result['updated']} updated, "
               f"{result['skipped']} skipped, {result['conflicts']} pending conflicts.")
    if result["errors"]:
        summary += f" {len(result['errors'])} error(s)."
    _push_alert({"status": "sync_complete", "messages": [summary],
                 "timestamp": result["timestamp"], "details": result})
    return result


def _get_scheduled_time() -> tuple[int, int] | None:
    return SYNC_SCHEDULE.get((TA_INCHARGE or "").lower().strip())


async def _daily_sync_scheduler():
    loop = asyncio.get_event_loop()
    last_sync_date = None
    while True:
        try:
            scheduled = _get_scheduled_time()
            if scheduled:
                now = datetime.now()
                h, m = scheduled
                if (now.hour, now.minute) >= (h, m) and now.date() != last_sync_date:
                    await loop.run_in_executor(None, _run_sync)
                    last_sync_date = now.date()
        except Exception as e:
            _push_alert({"status": "sync_error", "messages": [f"Scheduler error: {e}"],
                         "timestamp": datetime.now().isoformat()})
        await asyncio.sleep(30)  # check every 30s


# ---------------------------------------------------------------------------
# Lifespan — background sync scheduler only (no filesystem watcher anymore)
# ---------------------------------------------------------------------------
@contextlib.asynccontextmanager
async def _lifespan(app_instance):
    sync_task = asyncio.create_task(_daily_sync_scheduler())
    yield
    sync_task.cancel()
    try:
        await sync_task
    except asyncio.CancelledError:
        pass


# ---------------------------------------------------------------------------
# App + LLM
# ---------------------------------------------------------------------------
app = GreenNodeAgentBaseApp(
    lifespan=_lifespan,
    middleware=[
        Middleware(
            CORSMiddleware,
            allow_origins=["*"],
            allow_methods=["GET", "POST", "OPTIONS"],
            allow_headers=["*"],
        )
    ],
)

LLM_MODEL = os.environ.get("LLM_MODEL", "")
LLM_BASE_URL = os.environ.get("LLM_BASE_URL", "")
LLM_API_KEY = os.environ.get("LLM_API_KEY", "")
if not LLM_MODEL or not LLM_BASE_URL or not LLM_API_KEY:
    raise ValueError(
        "LLM_MODEL, LLM_BASE_URL, and LLM_API_KEY are required. "
        "Set them in .env or use /agentbase-llm to get a platform API key."
    )

llm = ChatOpenAI(model=LLM_MODEL, base_url=LLM_BASE_URL, api_key=LLM_API_KEY)


# ---------------------------------------------------------------------------
# LangGraph tools — Skill 1
# ---------------------------------------------------------------------------
@tool
def create_job_folder(folder_name: str) -> str:
    """Create a job folder record with standard recruitment sub-folders: LinkedIn, VNG Careers, Referral, TA Search, Others.

    Args:
        folder_name: Name of the job folder (e.g. 'ZDA - Data Scientist - 26-ZDA-3117').
    """
    if any(f["name"] == folder_name for f in job_folders):
        return f"Folder '{folder_name}' already exists."
    job_folders.append({
        "name": folder_name,
        "sub_folders": JOB_SUB_FOLDERS,
        "created_at": datetime.now().isoformat(),
    })
    _persist_folders()
    return f"Created '{folder_name}' with sub-folders: {', '.join(JOB_SUB_FOLDERS)}."


# ---------------------------------------------------------------------------
# LangGraph tools — Skill 2 (alerts + conflict resolution)
# ---------------------------------------------------------------------------
_RESOLVABLE_STATUSES = ("duplicate", "cross_role_duplicate", "sync_conflict_pending", "sync_cross_ta_pending")


def _format_conflict_alert(alert: dict) -> str | None:
    """Build a fully-substituted conflict block (including CONFLICT_DATA) directly from
    the alert's own fields, so the LLM never has to construct or retype the JSON itself."""
    status = alert.get("status")
    if status in ("duplicate", "cross_role_duplicate"):
        email = alert.get("email") or ""
        name = alert.get("candidate_name") or "Unknown"
        request_code = alert.get("request_code") or ""
        existing_job = alert.get("existing_job") or ""
        new_job = alert.get("new_job") or ""
        if status == "duplicate":
            issue = f"{name} already has a CV on file for this exact role ({request_code})."
        else:
            codes = ", ".join(alert.get("existing_roles") or [])
            issue = f"{name} is already in your pipeline for another role ({codes})."
        data = json.dumps({"email": email, "request_code": request_code, "type": status})
        return (f"⚠️ CONFLICT: {name} ({email})\n"
                f"Job: {request_code}\n"
                f"Already in: {existing_job}\n"
                f"New CV in: {new_job}\n"
                f"Issue: {issue}\n"
                f"CONFLICT_DATA:{data}")
    if status in ("sync_conflict_pending", "sync_cross_ta_pending"):
        email = alert.get("email") or ""
        name = alert.get("candidate_name") or "Unknown"
        request_code = alert.get("request_code") or ""
        team_ta = alert.get("team_ta") or ""
        issue = f"{team_ta} already has this candidate in the team database for {request_code}."
        data = json.dumps({"email": email, "request_code": request_code, "type": "sync_conflict"})
        return (f"⚠️ CONFLICT: {name} ({email})\n"
                f"Job: {request_code}\n"
                f"Issue: {issue}\n"
                f"CONFLICT_DATA:{data}")
    return None


@tool
def get_alerts() -> str:
    """Get all pending alerts: CV processing results, duplicates, sync outcomes, errors."""
    if not _alerts:
        return "No pending alerts."
    lines = []
    for i, alert in enumerate(_alerts[-20:], 1):
        status = alert.get("status", "")
        block = _format_conflict_alert(alert) if status in _RESOLVABLE_STATUSES else None
        if block:
            lines.append(f"[{i}] {alert.get('timestamp', '')[:19]}")
            lines.append(block)
        else:
            lines.append(f"[{i}] {alert.get('timestamp', '')[:19]} | {status} | {alert.get('file', 'system')}")
            for msg in alert.get("messages", []):
                lines.append(f"    {msg}")
    return "\n".join(lines)


@tool
def clear_alerts() -> str:
    """Clear all pending alerts."""
    _clear_alerts()
    return "All alerts cleared."


@tool
def resolve_duplicate(email: str, request_code: str, action: str) -> str:
    """Resolve a duplicate or cross-role duplicate candidate detected during CV processing.

    Args:
        email: The candidate's email address.
        request_code: The job request code of the NEW CV being processed.
        action: 'keep' to keep the existing record unchanged, 'overwrite' to replace the
            existing record with the new CV data (same-role duplicates only), or 'add' to
            add the new CV as a separate new row without touching the existing record.
    """
    pending = next((a for a in _alerts if a.get("status") in ("duplicate", "cross_role_duplicate")
                    and str(a.get("cv_fields", {}).get("email") or "").lower() == email.strip().lower()
                    and str(a.get("request_code") or "") == request_code.strip()), None)
    if not pending:
        return f"No pending duplicate for '{email}' / '{request_code}'."
    name = pending["cv_fields"].get("candidate_name") or "Unknown"
    is_cross_role = pending.get("status") == "cross_role_duplicate"
    action = action.lower()
    if action == "keep":
        _remove_alert(pending)
        return f"Kept existing record for {name}. No changes made."
    if action == "overwrite":
        if is_cross_role:
            return "Cannot overwrite: the existing record is for a different role. Use 'add' instead."
        row_data = _build_row_data(_CVFields(**pending["cv_fields"]), pending["folder_name"],
                                    pending["subfolder"], pending.get("referrer"), pending["filename"])
        ok = _overwrite_personal_row(pending["duplicate_id"], row_data)
        if not ok:
            return "Overwrite failed: original record no longer exists."
        _remove_alert(pending)
        return f"Overwrote existing record with updated data for {name}."
    if action == "add":
        row_data = _build_row_data(_CVFields(**pending["cv_fields"]), pending["folder_name"],
                                    pending["subfolder"], pending.get("referrer"), pending["filename"])
        new_row = _add_personal_row(row_data)
        _remove_alert(pending)
        return f"Added {name} as new row (No. {new_row['No']})."
    return "Invalid action. Use 'keep', 'overwrite', or 'add'."


@tool
def resolve_cross_role(email: str, new_request_code: str, action: str) -> str:
    """Resolve a cross-role duplicate: same candidate already in pipeline for a different role.

    Args:
        email: The candidate's email address.
        new_request_code: The new job request code being processed.
        action: 'add' to add the new row anyway, or 'skip' to not add.
    """
    pending = next((a for a in _alerts if a.get("status") == "cross_role_duplicate"
                    and str(a.get("cv_fields", {}).get("email") or "").lower() == email.strip().lower()
                    and str(a.get("request_code") or "") == new_request_code.strip()), None)
    if not pending:
        return f"No pending cross-role alert for '{email}' / '{new_request_code}'."
    name = pending["cv_fields"].get("candidate_name") or "Unknown"
    if action.lower() == "skip":
        _remove_alert(pending)
        return f"Skipped adding {name} for {new_request_code}. No changes made."
    if action.lower() == "add":
        row_data = _build_row_data(_CVFields(**pending["cv_fields"]), pending["folder_name"],
                                    pending["subfolder"], pending.get("referrer"), pending["filename"])
        new_row = _add_personal_row(row_data)
        _remove_alert(pending)
        return f"Added {name} for {new_request_code} at row (No. {new_row['No']})."
    return "Invalid action. Use 'add' or 'skip'."


@tool
def resolve_sync_conflict(email: str, request_code: str, action: str) -> str:
    """Resolve a pending sync conflict (different TA already has this candidate in the team database).

    Args:
        email: The candidate's email address.
        request_code: The job request code.
        action: 'skip' to leave the team DB unchanged, or 'add_new' to insert this row as a new entry.
    """
    pending = next((a for a in _alerts if a.get("status") in ("sync_conflict_pending", "sync_cross_ta_pending")
                    and str(a.get("email") or "").lower() == email.strip().lower()
                    and str(a.get("request_code") or "") == request_code.strip()), None)
    if not pending:
        return f"No pending sync conflict for '{email}' / '{request_code}'."
    candidate_name = pending.get("candidate_name") or "Unknown"
    if action.lower() == "skip":
        _remove_alert(pending)
        return f"Skipped syncing {candidate_name} ({email}) for {request_code}. Team DB unchanged."
    if action.lower() == "add_new":
        row_data = {k: v for k, v in pending["row_data"].items() if v is not None}
        new_row = _add_team_row(row_data)
        _remove_alert(pending)
        return f"Added {candidate_name} as new row (No. {new_row['No']}) in team database."
    return "Invalid action. Use 'skip' or 'add_new'."


# ---------------------------------------------------------------------------
# LangGraph tools — Skill 3
# ---------------------------------------------------------------------------
@tool
def run_sync_now() -> str:
    """Manually trigger the daily sync from personal database to team database right now."""
    result = _run_sync()
    lines = [f"Sync completed at {result['timestamp'][:19]}",
             f"  Added:     {result['added']}",
             f"  Updated:   {result['updated']}",
             f"  Skipped:   {result['skipped']} (no changes)",
             f"  Conflicts: {result['conflicts']}"]
    if result["errors"]:
        lines.append(f"  Errors: {'; '.join(result['errors'])}")
    return "\n".join(lines)


@tool
def get_sync_status() -> str:
    """Show when the last sync ran and when the next one is scheduled for this TA."""
    last = (_sync_state.get("last_sync") or "Never")[:19]
    scheduled = _get_scheduled_time()
    sched_str = f"{scheduled[0]:02d}:{scheduled[1]:02d} daily" if scheduled else f"Not scheduled (TA '{TA_INCHARGE}' not in schedule)"
    rows_tracked = len(_sync_state.get("rows", {}))
    return (f"TA Incharge:    {TA_INCHARGE}\n"
            f"Scheduled sync: {sched_str}\n"
            f"Last sync:      {last}\n"
            f"Rows tracked:   {rows_tracked}")


# ---------------------------------------------------------------------------
# LangGraph graph
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """You are a Recruitment Agent.

Skills:
1. create_job_folder     - create a job folder record with standard sub-folders (LinkedIn, VNG Careers, Referral, TA Search, Others)
2. get_alerts             - check pending alerts: CV results, duplicates, sync conflicts, errors
3. clear_alerts           - dismiss resolved alerts
4. resolve_duplicate      - same email (same or different job code): keep, overwrite (same job code only), or add as new
5. resolve_cross_role     - same email + different job code: add or skip (legacy path, prefer resolve_duplicate)
6. resolve_sync_conflict  - sync blocked by a different TA already having this candidate: add new row or skip
7. run_sync_now           - manually trigger personal -> team database sync
8. get_sync_status        - show last sync time and next scheduled sync

CVs are uploaded and parsed through the Upload CV wizard in the chat UI — you do not have a
tool to process a CV file yourself; only react to the alerts that wizard produces.

Rules:
- Never fill Stage, Status, Note, Reason for failure, or salary fields — recruiter only
  (exception: auto-write "Email missing - please fill in manually" in Note when email is missing)
- Never delete existing rows; only add or (on explicit recruiter instruction) overwrite
- Missing CV fields: leave blank, alert recruiter — do not guess
- Referrer: blank unless source = Referral
- Sync: only database rows, never copy CV files

Sync workflow:
- After run_sync_now() completes, ALWAYS call get_alerts() immediately as your next step.
- Report in plain language: "Sync complete: X added, Y updated, Z unchanged, N conflicts need attention."
- If conflicts exist, show them immediately — never ask "would you like to check alerts?".
- get_alerts() already formats each conflict in full, including the CONFLICT_DATA line with
  the real email/request_code/type values filled in. Copy every conflict block from
  get_alerts() into your reply EXACTLY as returned, character for character — including the
  CONFLICT_DATA line. Never retype, reformat, summarize, or recompute the CONFLICT_DATA line
  yourself; never write a placeholder like "[request_code]" — always use get_alerts()'s own
  text verbatim.
- After listing all conflicts, say: "Use the buttons above to resolve each conflict."

Proactive suggestions (add 1-2 sentences after completing each task):
- After create_job_folder -> suggest uploading CVs to the new folder.
- After run_sync_now -> suggest checking alerts if any conflicts were flagged.
- After resolving a conflict -> suggest viewing the Database page to confirm the change.
Keep suggestions brief (1 sentence each), natural, and only when they add value. Do not repeat the same suggestion twice in a row."""


class State(TypedDict):
    messages: Annotated[list, add_messages]


tools = [
    create_job_folder,
    get_alerts, clear_alerts,
    resolve_duplicate, resolve_cross_role, resolve_sync_conflict,
    run_sync_now, get_sync_status,
]
llm_with_tools = llm.bind_tools(tools)


def chatbot(state: State) -> dict:
    messages = [SystemMessage(content=SYSTEM_PROMPT)] + state["messages"]
    return {"messages": [llm_with_tools.invoke(messages)]}


graph_builder = StateGraph(State)
graph_builder.add_node("chatbot", chatbot)
graph_builder.add_node("tools", ToolNode(tools))
graph_builder.add_edge(START, "chatbot")
graph_builder.add_conditional_edges("chatbot", tools_condition)
graph_builder.add_edge("tools", "chatbot")
graph = graph_builder.compile()


# ---------------------------------------------------------------------------
# Ensure CONFLICT_DATA lines survive even if the LLM paraphrases its reply
# ---------------------------------------------------------------------------
_CONFLICT_DATA_LINE_RE = re.compile(r'CONFLICT_DATA:\{[^\n\r]+\}')


def _ensure_conflict_markers(messages: list, final_text: str) -> str:
    extra_lines: list[str] = []
    for m in messages:
        if not isinstance(m, ToolMessage):
            continue
        content = m.content if isinstance(m.content, str) else str(m.content)
        for match in _CONFLICT_DATA_LINE_RE.finditer(content):
            line = match.group(0)
            if line not in final_text and line not in extra_lines:
                extra_lines.append(line)
    if not extra_lines:
        return final_text
    return final_text.rstrip() + "\n" + "\n".join(extra_lines)


# ---------------------------------------------------------------------------
# HTTP handlers — CV upload wizard (Skill 2)
# ---------------------------------------------------------------------------
_staged_files: dict[str, dict] = {}
_staged_lock = threading.Lock()
_parse_jobs: dict[str, dict] = {}


async def handle_save_cv(request: Request) -> JSONResponse:
    """Phase 1 — stage an uploaded CV in memory immediately; no AI parsing yet."""
    try:
        form = await request.form()
    except Exception as exc:
        return JSONResponse({"status": "error", "response": f"Could not parse form: {exc}"}, status_code=400)

    folder    = str(form.get("folder")    or "").strip()
    subfolder = str(form.get("subfolder") or "").strip()
    referrer  = str(form.get("referrer")  or "").strip() or None
    upload    = form.get("file")

    if not folder or not subfolder or upload is None:
        return JSONResponse(
            {"status": "error", "response": "Fields 'folder', 'subfolder', and 'file' are required."},
            status_code=400,
        )
    if subfolder not in JOB_SUB_FOLDERS:
        return JSONResponse(
            {"status": "error", "response": f"'subfolder' must be one of: {', '.join(JOB_SUB_FOLDERS)}"},
            status_code=400,
        )

    raw_name = Path(upload.filename).name
    ext = Path(raw_name).suffix.lower()
    if ext not in _ALLOWED_EXT:
        return JSONResponse(
            {"status": "error", "response": f"File type '{ext}' not supported. Use PDF or DOCX."},
            status_code=400,
        )

    contents = await upload.read()
    file_id = uuid.uuid4().hex[:16]
    with _staged_lock:
        _staged_files[file_id] = {
            "filename": raw_name, "data": contents,
            "folder": folder, "subfolder": subfolder, "referrer": referrer,
        }

    if not any(f["name"] == folder for f in job_folders):
        job_folders.append({"name": folder, "sub_folders": JOB_SUB_FOLDERS,
                             "created_at": datetime.now().isoformat()})
        _persist_folders()

    return JSONResponse({
        "status":    "saved",
        "name":      raw_name,
        "saved_to":  file_id,
        "folder":    folder,
        "subfolder": subfolder,
        "timestamp": datetime.now().isoformat(),
    })


def _parse_one(file_id: str, timeout: float = 120.0) -> dict:
    staged = _staged_files.get(file_id)
    if not staged:
        return {"status": "error", "messages": ["File no longer available (already processed or expired)."]}

    result: dict = {}
    exc_holder: list = []

    def _run() -> None:
        try:
            result.update(_process_cv(staged["filename"], staged["data"], staged["folder"],
                                       staged["subfolder"], staged.get("referrer")))
        except Exception as e:  # noqa: BLE001
            exc_holder.append(e)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    t.join(timeout=timeout)
    if t.is_alive():
        return {"status": "timeout", "messages": [f"Parsing timed out after {int(timeout)}s."]}
    if exc_holder:
        raise exc_holder[0]
    return result or {"status": "error", "messages": ["No result returned."]}


def _parse_worker(job_id: str, file_ids: list[str]) -> None:
    job = _parse_jobs[job_id]

    def _run_one(file_id: str) -> dict:
        staged = _staged_files.get(file_id)
        name = staged["filename"] if staged else file_id
        job["current"] = name
        try:
            result = _parse_one(file_id)
            return {"name": name, **result, "path": file_id}
        except Exception as e:  # noqa: BLE001
            return {"name": name, "status": "error", "messages": [str(e)], "path": file_id}
        finally:
            with _staged_lock:
                _staged_files.pop(file_id, None)

    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {pool.submit(_run_one, fid): fid for fid in file_ids}
        for fut in as_completed(futures):
            result = fut.result()
            status = result.get("status", "error")
            if status in ("success", "duplicate", "cross_role_duplicate"):
                job["done"] += 1
            else:
                job["failed"] += 1
            job["results"].append(result)
            job["progress"] += 1

    job["status"] = "complete" if job["failed"] == 0 else ("partial" if job["done"] > 0 else "failed")
    job["finished_at"] = datetime.now().isoformat()


async def handle_start_parse(request: Request) -> JSONResponse:
    """Phase 2 — kick off background parsing for previously staged file ids."""
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"status": "error", "response": "Request body must be valid JSON."}, status_code=400)

    files: list[str] = body.get("files", [])
    if not files:
        return JSONResponse({"status": "error", "response": "No files to parse."}, status_code=400)

    job_id = uuid.uuid4().hex[:12]
    _parse_jobs[job_id] = {
        "status":      "running",
        "total":       len(files),
        "progress":    0,
        "done":        0,
        "failed":      0,
        "current":     "",
        "results":     [],
        "started_at":  datetime.now().isoformat(),
        "finished_at": None,
    }

    threading.Thread(target=_parse_worker, args=(job_id, files), daemon=True).start()
    return JSONResponse({"job_id": job_id, "total": len(files), "status": "running"})


async def handle_parse_status(request: Request) -> JSONResponse:
    job_id = request.path_params.get("job_id", "")
    job = _parse_jobs.get(job_id)
    if not job:
        return JSONResponse({"status": "error", "response": f"Unknown job_id: {job_id!r}"}, status_code=404)
    return JSONResponse({"job_id": job_id, **job})


async def handle_list_folders(request: Request) -> JSONResponse:
    return JSONResponse({"folders": [f["name"] for f in job_folders]})


# ---------------------------------------------------------------------------
# HTTP handlers — Database portal
# ---------------------------------------------------------------------------
async def handle_database(request: Request) -> JSONResponse:
    return JSONResponse({
        "personal": [_public_row(r) for r in personal_db],
        "team":     [_public_row(r) for r in team_db],
    })


async def handle_download_excel(request: Request) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Database"
    for col, h in enumerate(DB_COLUMNS, 1):
        ws.cell(row=1, column=col, value=h)
    # Row 2 left blank (annotation row); data starts row 3
    for r_idx, row in enumerate(personal_db, start=3):
        for col, h in enumerate(DB_COLUMNS, 1):
            ws.cell(row=r_idx, column=col, value=row.get(h))
    buf = io.BytesIO()
    wb.save(buf)
    filename = f"personal_db_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


app.router.routes.extend([
    Route("/save-cv",               handle_save_cv,       methods=["POST"]),
    Route("/start-parse",           handle_start_parse,   methods=["POST"]),
    Route("/parse-status/{job_id}", handle_parse_status,  methods=["GET"]),
    Route("/list-folders",          handle_list_folders,  methods=["GET"]),
    Route("/database",              handle_database,       methods=["GET"]),
    Route("/download-excel",        handle_download_excel, methods=["GET"]),
])


# ---------------------------------------------------------------------------
# Entrypoint & health check
# ---------------------------------------------------------------------------
@app.entrypoint
def handler(payload: dict, context: RequestContext) -> dict:
    message = payload.get("message", "Hello")
    result = graph.invoke({"messages": [("user", message)]})
    final_text = _ensure_conflict_markers(result["messages"], result["messages"][-1].content)
    return {
        "status": "success",
        "response": final_text,
        "timestamp": datetime.now().isoformat(),
    }


@app.ping
def health_check() -> PingStatus:
    return PingStatus.HEALTHY


if __name__ == "__main__":
    app.run(port=8080, host="0.0.0.0")
