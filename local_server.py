"""
Local HTTP server for recruiter-agent.

Usage:
  python local_server.py
  ngrok http 8080

Endpoints:
  POST /invocations            {"message": "..."}              -> agent chat response
  POST /save-cv                multipart: folder, subfolder, file -> save file, no AI
  POST /start-parse            {"files": [...paths]}           -> start background AI parse
  GET  /parse-status/{job_id}                                  -> poll parse job status
  GET  /list-folders                                           -> list job folders on disk
  GET  /health                                                 -> {"status": "healthy"}

Standalone by design — does NOT import main.py because GreenNodeAgentBaseApp
crashes on Windows cp1252 terminals at import time (rich printing ✓).
main.py is kept for AgentBase cloud deployment only.
"""

import asyncio
import json
import os
import re
import subprocess
import threading
import unicodedata
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Annotated, Optional, TypedDict

import uvicorn
from dotenv import load_dotenv
from langchain_core.messages import SystemMessage, ToolMessage
from langchain_core.tools import tool
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, START
from langgraph.graph.message import add_messages
from langgraph.prebuilt import ToolNode, tools_condition
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel
from starlette.applications import Starlette
from starlette.middleware import Middleware
from starlette.middleware.cors import CORSMiddleware
from starlette.requests import Request
from starlette.responses import JSONResponse
from starlette.routing import Route

load_dotenv()

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
JOBS_BASE_DIR = os.environ.get("JOBS_BASE_DIR", os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH    = os.environ.get("EXCEL_PATH", "")
TA_INCHARGE   = os.environ.get("TA_INCHARGE") or os.environ.get("USERNAME", "Unknown")
PROCESSED_FILES_LOG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_cvs.json")

JOB_SUB_FOLDERS = ["LinkedIn", "VNG Careers", "Referral", "TA Search", "Others"]


def _safe_join(base_dir: str, *parts: str) -> Path:
    """Join parts onto base_dir, raising ValueError if the result would escape base_dir."""
    base = Path(base_dir).resolve()
    target = (base / Path(*parts)).resolve()
    if target != base and not target.is_relative_to(base):
        raise ValueError(f"Path escapes base directory: {Path(*parts)}")
    return target

DB_COLUMNS = [
    "No", "Request code", "Candidate name", "Processed Team", "Processed Position",
    "Entry date", "Source", "Referrer", "Latest company", "Latest position",
    "Email", "Phone", "Stage", "Status", "Note", "Reason for failure/withdrawal",
    "Last drawn salary", "Expected salary (Monthly Gross)", "TA Incharge", "Profile",
]

TEAM_EXCEL_PATH = os.environ.get("TEAM_EXCEL_PATH", "")
SYNC_STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sync_state.json")
CONFLICT_SHEET  = "Conflict Log"

SYNC_SCHEDULE: dict[str, tuple[int, int]] = {
    "trangptt12": (7, 30),
    "hautt2":     (8,  0),
    "huyenplt":   (8, 30),
    "nhihm":      (9,  0),
}

SYNC_COLUMNS = [c for c in DB_COLUMNS if c not in ("No", "Profile")]

CONFLICT_LOG_COLUMNS = [
    "Timestamp", "Email", "Request code", "Candidate name",
    "Field", "My value", "Team value", "My TA", "Team TA",
]

_DATA_DIR    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
ALERTS_PATH  = os.path.join(_DATA_DIR, "alerts.json")

os.makedirs(_DATA_DIR, exist_ok=True)

def _load_alerts() -> list[dict]:
    try:
        with open(ALERTS_PATH, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def _save_alerts() -> None:
    with open(ALERTS_PATH, "w", encoding="utf-8") as f:
        json.dump(_alerts, f, ensure_ascii=False, indent=2, default=str)

def _alert_push(alert: dict) -> None:
    _alerts.append(alert)
    _save_alerts()

def _alert_remove(item: dict) -> None:
    _alerts.remove(item)
    _save_alerts()

def _alert_clear() -> None:
    _alerts.clear()
    _save_alerts()

_alerts: list[dict] = _load_alerts()

# ---------------------------------------------------------------------------
# LLM
# ---------------------------------------------------------------------------
LLM_MODEL   = os.environ.get("LLM_MODEL", "")
LLM_BASE_URL = os.environ.get("LLM_BASE_URL", "")
LLM_API_KEY  = os.environ.get("LLM_API_KEY", "")
if not LLM_MODEL or not LLM_BASE_URL or not LLM_API_KEY:
    raise ValueError(
        "LLM_MODEL, LLM_BASE_URL, and LLM_API_KEY are required. "
        "Check your .env file."
    )
llm = ChatOpenAI(model=LLM_MODEL, base_url=LLM_BASE_URL, api_key=LLM_API_KEY)

# ---------------------------------------------------------------------------
# Shared Excel helpers
# ---------------------------------------------------------------------------
def _open_wb(path: str, read_only: bool = False):
    keep_vba = Path(path).suffix.lower() == ".xlsm"
    return load_workbook(path, read_only=read_only, keep_vba=keep_vba)

def _get_headers(ws) -> list[str]:
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c).strip() if c is not None else "" for c in row]
    return []

# ---------------------------------------------------------------------------
# Excel in-memory cache  (loaded at startup, updated on every write)
# ---------------------------------------------------------------------------
_excel_cache:   dict[str, list[dict]] = {}   # email_lower → list of row metadata
_excel_lock     = threading.Lock()
_excel_pending: list[dict] = []              # rows staged for batch flush
_excel_next_row: int       = 3               # next assignable row number in Excel

def _init_excel_cache() -> None:
    """Load existing Excel rows into _excel_cache for O(1) duplicate lookups."""
    global _excel_cache, _excel_next_row
    _excel_cache = {}
    _excel_next_row = 3
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        return
    wb = _open_wb(EXCEL_PATH, read_only=True)
    if "Database" not in wb.sheetnames:
        wb.close(); return
    ws = wb["Database"]
    headers = None; max_row = 2
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if headers is None:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue
        if not any(c is not None for c in row):
            continue
        row_dict = dict(zip(headers, row))
        email_l = str(row_dict.get("Email") or "").strip().lower()
        if email_l:
            _excel_cache.setdefault(email_l, []).append({
                "request_code":       str(row_dict.get("Request code") or "").strip(),
                "candidate_name":     row_dict.get("Candidate name"),
                "ta_incharge":        str(row_dict.get("TA Incharge") or "").strip().lower(),
                "processed_team":     str(row_dict.get("Processed Team") or "").strip(),
                "processed_position": str(row_dict.get("Processed Position") or "").strip(),
                "row_num":            row_idx,
            })
        max_row = row_idx
    _excel_next_row = max_row + 1
    wb.close()

def _check_duplicate_cache(email: str, request_code: str) -> dict | None:
    email_l = email.strip().lower()
    rc      = request_code.strip()
    with _excel_lock:
        for entry in _excel_cache.get(email_l, []):
            if entry["request_code"] == rc:
                return {
                    "row_number":         entry["row_num"],
                    "candidate_name":     entry["candidate_name"],
                    "processed_team":     entry.get("processed_team", ""),
                    "processed_position": entry.get("processed_position", ""),
                }
    return None

def _check_cross_role_cache(email: str, exclude_request_code: str) -> list[dict]:
    email_l = email.strip().lower()
    exc_rc  = exclude_request_code.strip()
    ta_l    = (TA_INCHARGE or "").strip().lower()
    found   = []
    with _excel_lock:
        for entry in _excel_cache.get(email_l, []):
            if entry["request_code"] != exc_rc and entry["ta_incharge"] == ta_l:
                found.append({
                    "row_number":         entry["row_num"],
                    "candidate_name":     entry["candidate_name"],
                    "request_code":       entry["request_code"],
                    "processed_team":     entry.get("processed_team", ""),
                    "processed_position": entry.get("processed_position", ""),
                })
    return found

def _cache_add_row(row_data: dict) -> int:
    """Stage a new row in memory. Flush to disk later with _flush_excel_pending()."""
    global _excel_next_row
    email_l = str(row_data.get("Email") or "").strip().lower()
    with _excel_lock:
        row_num = _excel_next_row
        _excel_next_row += 1
        if email_l:
            _excel_cache.setdefault(email_l, []).append({
                "request_code":       str(row_data.get("Request code") or "").strip(),
                "candidate_name":     row_data.get("Candidate name"),
                "ta_incharge":        str(row_data.get("TA Incharge") or "").strip().lower(),
                "processed_team":     str(row_data.get("Processed Team") or "").strip(),
                "processed_position": str(row_data.get("Processed Position") or "").strip(),
                "row_num":            row_num,
            })
        _excel_pending.append({"__row_num__": row_num, **row_data})
    return row_num

def _flush_excel_pending() -> int:
    """Write all staged rows to Excel in a single wb.save(). Returns count written."""
    with _excel_lock:
        if not _excel_pending:
            return 0
        pending = list(_excel_pending)
        _excel_pending.clear()
    if not EXCEL_PATH:
        return 0
    wb = _open_wb(EXCEL_PATH)
    if "Database" not in wb.sheetnames:
        wb.create_sheet("Database")
    ws = wb["Database"]
    headers = _get_headers(ws)
    if not any(headers):
        for col, h in enumerate(DB_COLUMNS, 1):
            ws.cell(row=1, column=col, value=h)
        headers = DB_COLUMNS
    for row_data in sorted(pending, key=lambda r: r.get("__row_num__", 0)):
        row_num = row_data["__row_num__"]
        if "No" in headers:
            ws.cell(row=row_num, column=headers.index("No") + 1, value=row_num - 2)
        for col_name, value in row_data.items():
            if col_name in ("__row_num__", "Profile"):
                continue
            if col_name in headers:
                ws.cell(row=row_num, column=headers.index(col_name) + 1, value=value)
        if "Profile" in headers and row_data.get("Profile"):
            fp   = str(row_data["Profile"])
            disp = str(row_data.get("Candidate name") or os.path.basename(fp))
            url  = "file:///" + fp.replace("\\", "/")
            cell = ws.cell(row=row_num, column=headers.index("Profile") + 1, value=disp)
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
    wb.save(EXCEL_PATH)
    wb.close()
    return len(pending)

# ---------------------------------------------------------------------------
# CV text extraction
# ---------------------------------------------------------------------------
def _resolve_path(file_path: str) -> str:
    file_path = os.path.normpath(file_path)
    if os.path.exists(file_path):
        return file_path
    nfc = unicodedata.normalize("NFC", file_path)
    if os.path.exists(nfc):
        return nfc
    parent, name = os.path.dirname(file_path), os.path.basename(file_path)
    name_nfc = unicodedata.normalize("NFC", name).lower()
    if os.path.isdir(parent):
        for fname in os.listdir(parent):
            if unicodedata.normalize("NFC", fname).lower() == name_nfc:
                return os.path.join(parent, fname)
    return file_path

def _extract_pdf_text(file_path: str) -> str:
    import pypdf
    with open(file_path, "rb") as f:
        reader = pypdf.PdfReader(f)
        return "\n".join(page.extract_text() or "" for page in reader.pages)

def _extract_docx_text(file_path: str) -> str:
    from docx import Document
    with open(file_path, "rb") as f:
        doc = Document(f)
    return "\n".join(p.text for p in doc.paragraphs)

def _extract_cv_text(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":  return _extract_pdf_text(file_path)
    if ext == ".docx": return _extract_docx_text(file_path)
    raise ValueError(f"Unsupported file type '{ext}'. Only PDF and DOCX are supported.")

# ---------------------------------------------------------------------------
# LLM CV field extraction
# ---------------------------------------------------------------------------
class _CVFields(BaseModel):
    full_name:       Optional[str] = None
    email:           Optional[str] = None
    phone:           Optional[str] = None
    current_company: Optional[str] = None
    current_title:   Optional[str] = None

def _parse_cv_with_llm(text: str) -> _CVFields:
    structured_llm = llm.with_structured_output(_CVFields)
    return structured_llm.invoke(
        "Extract CV fields. Return null for missing data. Reply with JSON only. No preamble.\n"
        f"CV:\n{text[:3000]}"
    )

# ---------------------------------------------------------------------------
# Folder path parsing
# ---------------------------------------------------------------------------
def _parse_folder_path(file_path: str) -> dict:
    parts = Path(os.path.normpath(file_path)).parts
    source = job_folder = None
    for i, part in enumerate(parts):
        if part in JOB_SUB_FOLDERS and i > 0:
            source = part
            job_folder = parts[i - 1]
            break
    if not source or not job_folder:
        return {}
    segments = job_folder.split(" - ", 2)
    if len(segments) < 3:
        return {}
    return {
        "processed_team":     segments[0].strip(),
        "processed_position": segments[1].strip(),
        "request_code":       segments[2].strip(),
        "source":             source,
    }

# ---------------------------------------------------------------------------
# Personal DB Excel operations
# ---------------------------------------------------------------------------
def _check_duplicate(email: str, request_code: str) -> dict | None:
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        return None
    wb = _open_wb(EXCEL_PATH, read_only=True)
    if "Database" not in wb.sheetnames:
        wb.close(); return None
    ws = wb["Database"]
    headers = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if headers is None:
            headers = [str(c).strip() if c is not None else "" for c in row]; continue
        row_dict = dict(zip(headers, row))
        if (str(row_dict.get("Email") or "").strip().lower() == email.strip().lower()
                and str(row_dict.get("Request code") or "").strip() == request_code.strip()):
            wb.close()
            return {"row_number": row_idx, "candidate_name": row_dict.get("Candidate name")}
    wb.close(); return None

def _check_cross_role_duplicate(email: str, exclude_request_code: str) -> list[dict]:
    if not email or not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        return []
    wb = _open_wb(EXCEL_PATH, read_only=True)
    if "Database" not in wb.sheetnames:
        wb.close(); return []
    ws = wb["Database"]; headers = None; found = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if headers is None:
            headers = [str(c).strip() if c is not None else "" for c in row]; continue
        row_dict = dict(zip(headers, row))
        row_email = str(row_dict.get("Email") or "").strip().lower()
        row_code  = str(row_dict.get("Request code") or "").strip()
        row_ta    = str(row_dict.get("TA Incharge") or "").strip().lower()
        if (row_email == email.strip().lower()
                and row_code != exclude_request_code.strip()
                and row_ta == (TA_INCHARGE or "").lower().strip()):
            found.append({"row_number": row_idx, "candidate_name": row_dict.get("Candidate name"), "request_code": row_code})
    wb.close(); return found

def _add_row(row_data: dict) -> int:
    """Write one row directly to Excel (used by agent tools). Flushes pending batch first."""
    global _excel_next_row
    _flush_excel_pending()   # ensure file is consistent before we open it
    if not EXCEL_PATH:
        raise ValueError("EXCEL_PATH is not configured in .env")
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")
    with _excel_lock:
        row_num = _excel_next_row
        _excel_next_row += 1
    wb = _open_wb(EXCEL_PATH)
    if "Database" not in wb.sheetnames:
        wb.create_sheet("Database")
    ws = wb["Database"]
    headers = _get_headers(ws)
    if not any(headers):
        for col, h in enumerate(DB_COLUMNS, 1):
            ws.cell(row=1, column=col, value=h)
        headers = DB_COLUMNS
    if "No" in headers:
        ws.cell(row=row_num, column=headers.index("No") + 1, value=row_num - 2)
    for col_name, value in row_data.items():
        if col_name == "Profile": continue
        if col_name in headers:
            ws.cell(row=row_num, column=headers.index(col_name) + 1, value=value)
    if "Profile" in headers and row_data.get("Profile"):
        fp   = str(row_data["Profile"])
        disp = str(row_data.get("Candidate name") or os.path.basename(fp))
        url  = "file:///" + fp.replace("\\", "/")
        cell = ws.cell(row=row_num, column=headers.index("Profile") + 1, value=disp)
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")
    wb.save(EXCEL_PATH); wb.close()
    email_l = str(row_data.get("Email") or "").strip().lower()
    if email_l:
        with _excel_lock:
            _excel_cache.setdefault(email_l, []).append({
                "request_code":       str(row_data.get("Request code") or "").strip(),
                "candidate_name":     row_data.get("Candidate name"),
                "ta_incharge":        str(row_data.get("TA Incharge") or "").strip().lower(),
                "processed_team":     str(row_data.get("Processed Team") or "").strip(),
                "processed_position": str(row_data.get("Processed Position") or "").strip(),
                "row_num":            row_num,
            })
    return row_num

def _overwrite_row(row_number: int, row_data: dict):
    _flush_excel_pending()   # ensure file is consistent before opening
    wb = _open_wb(EXCEL_PATH); ws = wb["Database"]
    headers = _get_headers(ws)
    for col_name, value in row_data.items():
        if col_name == "Profile": continue
        if col_name in headers:
            ws.cell(row=row_number, column=headers.index(col_name) + 1, value=value)
    if "Profile" in headers and row_data.get("Profile"):
        fp   = str(row_data["Profile"])
        disp = str(row_data.get("Candidate name") or os.path.basename(fp))
        url  = "file:///" + fp.replace("\\", "/")
        cell = ws.cell(row=row_number, column=headers.index("Profile") + 1, value=disp)
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single")
    wb.save(EXCEL_PATH); wb.close()
    email_l = str(row_data.get("Email") or "").strip().lower()
    if email_l:
        with _excel_lock:
            for entry in _excel_cache.get(email_l, []):
                if entry["row_num"] == row_number:
                    entry["candidate_name"] = row_data.get("Candidate name")
                    entry["ta_incharge"]    = str(row_data.get("TA Incharge") or "").strip().lower()
                    break

# ---------------------------------------------------------------------------
# Processed-files tracking
# ---------------------------------------------------------------------------
def _load_processed() -> set:
    if os.path.exists(PROCESSED_FILES_LOG):
        with open(PROCESSED_FILES_LOG) as f:
            return set(json.load(f))
    return set()

def _mark_processed(file_path: str):
    processed = _load_processed()
    processed.add(os.path.normpath(file_path))
    with open(PROCESSED_FILES_LOG, "w") as f:
        json.dump(list(processed), f, indent=2)

# ---------------------------------------------------------------------------
# Core CV processing pipeline
# ---------------------------------------------------------------------------
def _build_row_data(cv: _CVFields, folder: dict, file_path: str) -> dict:
    return {
        "Request code":       folder.get("request_code"),
        "Candidate name":     cv.full_name,
        "Processed Team":     folder.get("processed_team"),
        "Processed Position": folder.get("processed_position"),
        "Entry date":         datetime.now().strftime("%Y-%m-%d"),
        "Source":             folder.get("source"),
        "Referrer":           None,
        "Latest company":     cv.current_company,
        "Latest position":    cv.current_title,
        "Email":              cv.email,
        "Phone":              cv.phone,
        "TA Incharge":        TA_INCHARGE,
        "Profile":            os.path.normpath(file_path),
    }

def _process_cv(file_path: str) -> dict:
    file_path = _resolve_path(file_path)
    result: dict = {"file": os.path.basename(file_path), "status": None, "messages": []}
    try:
        text = _extract_cv_text(file_path)
    except Exception as e:
        result.update(status="error", messages=[f"Cannot extract text: {e}"])
        _alert_push({**result, "timestamp": datetime.now().isoformat()}); return result
    if not text.strip():
        result.update(status="error", messages=["CV appears empty or image-only."])
        _alert_push({**result, "timestamp": datetime.now().isoformat()}); return result
    folder = _parse_folder_path(file_path)
    if not folder:
        result.update(status="error", messages=["Could not parse folder path. Expected: [Dept] - [Position] - [JobCode]/[SubFolder]/file"])
        _alert_push({**result, "timestamp": datetime.now().isoformat()}); return result
    try:
        cv = _parse_cv_with_llm(text)
    except Exception as e:
        result.update(status="error", messages=[f"LLM extraction failed: {e}"])
        _alert_push({**result, "timestamp": datetime.now().isoformat()}); return result
    missing = [label for label, val in [
        ("Candidate name", cv.full_name), ("Email", cv.email), ("Phone", cv.phone),
        ("Latest company", cv.current_company), ("Latest position", cv.current_title),
    ] if not val]
    if missing:
        result["messages"].append(f"Warning - Missing fields (left blank): {', '.join(missing)}")
    new_job = " - ".join(filter(None, [folder.get("processed_team"), folder.get("processed_position"), folder.get("request_code")]))
    if cv.email:
        dup = _check_duplicate_cache(cv.email, folder["request_code"])
        if dup:
            existing_job = " - ".join(filter(None, [dup.get("processed_team"), dup.get("processed_position"), folder["request_code"]]))
            result.update(status="duplicate", messages=result["messages"] + [
                f"DUPLICATE: {cv.full_name or 'Unknown'} ({cv.email}) already exists "
                f"in row {dup['row_number']} for {folder['request_code']}. "
                "Use resolve_duplicate to keep, overwrite, or add as new."
            ], email=cv.email, candidate_name=cv.full_name, request_code=folder["request_code"],
               new_job=new_job, new_source=folder.get("source"), existing_job=existing_job)
            _alert_push({**result, "timestamp": datetime.now().isoformat(),
                "file_path": os.path.normpath(file_path), "cv_fields": cv.model_dump(),
                "folder_info": folder, "duplicate_row": dup["row_number"]})
            _mark_processed(file_path); return result
        cross_role = _check_cross_role_cache(cv.email, folder["request_code"])
        if cross_role:
            codes = ", ".join(r["request_code"] for r in cross_role)
            first = cross_role[0]
            existing_job = " - ".join(filter(None, [first.get("processed_team"), first.get("processed_position"), first.get("request_code")]))
            result.update(status="cross_role_duplicate", messages=result["messages"] + [
                f"CROSS-ROLE: {cv.full_name or 'This candidate'} ({cv.email}) is already "
                f"in your pipeline for another role ({codes}). "
                "Use resolve_duplicate to keep or add as new."
            ], email=cv.email, candidate_name=cv.full_name, request_code=folder["request_code"],
               new_job=new_job, new_source=folder.get("source"), existing_job=existing_job)
            _alert_push({**result, "timestamp": datetime.now().isoformat(),
                "file_path": os.path.normpath(file_path), "cv_fields": cv.model_dump(),
                "folder_info": folder, "existing_roles": [r["request_code"] for r in cross_role]})
            _mark_processed(file_path); return result
    row_data = _build_row_data(cv, folder, file_path)
    if not cv.email:
        row_data["Note"] = "Email missing - please fill in manually"
    if EXCEL_PATH:
        try:
            new_row = _cache_add_row(row_data)   # stage in memory; flushed after batch
            result.update(status="success")
            result["messages"].append(f"Added {cv.full_name or 'candidate'} to database at row {new_row}.")
        except Exception as e:
            result.update(status="error")
            result["messages"].append(f"Excel write failed: {e}")
    else:
        result.update(status="no_excel")
        result["messages"].append("EXCEL_PATH not set - candidate NOT written to database.")
    _alert_push({**result, "timestamp": datetime.now().isoformat()})
    _mark_processed(file_path)
    return result

# ---------------------------------------------------------------------------
# Skill 3 — Sync helpers
# ---------------------------------------------------------------------------
def _row_key(row: dict) -> str:
    email = str(row.get("Email") or "").strip().lower()
    code  = str(row.get("Request code") or "").strip()
    return f"{email}|{code}"

def _load_sync_state() -> dict:
    if os.path.exists(SYNC_STATE_FILE):
        with open(SYNC_STATE_FILE) as f: return json.load(f)
    return {"last_sync": None, "rows": {}}

def _save_sync_state(personal_rows: list[dict]):
    state = {"last_sync": datetime.now().isoformat(),
             "rows": {_row_key(r): {k: str(v) if v is not None else None for k, v in r.items()}
                      for r in personal_rows if _row_key(r) not in ("|", "")}}
    with open(SYNC_STATE_FILE, "w") as f: json.dump(state, f, indent=2)

def _row_changed_vs_snapshot(current: dict, snapshot: dict) -> bool:
    return any(str(current.get(c) or "").strip() != str(snapshot.get(c) or "").strip() for c in SYNC_COLUMNS)

def _read_personal_db() -> list[dict]:
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH): return []
    wb = _open_wb(EXCEL_PATH, read_only=True)
    if "Database" not in wb.sheetnames: wb.close(); return []
    ws = wb["Database"]; headers = None; rows = []
    for row in ws.iter_rows(values_only=True):
        if headers is None: headers = [str(c).strip() if c is not None else "" for c in row]; continue
        if not any(c is not None for c in row): continue
        rows.append(dict(zip(headers, row)))
    wb.close(); return rows

def _read_team_db() -> dict[str, tuple[int, dict]]:
    if not TEAM_EXCEL_PATH or not os.path.exists(TEAM_EXCEL_PATH): return {}
    wb = _open_wb(TEAM_EXCEL_PATH, read_only=True)
    if "Database" not in wb.sheetnames: wb.close(); return {}
    ws = wb["Database"]; lookup: dict[str, tuple[int, dict]] = {}; headers = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if headers is None: headers = [str(c).strip() if c is not None else "" for c in row]; continue
        if not any(c is not None for c in row): continue
        row_dict = dict(zip(headers, row)); key = _row_key(row_dict)
        if key and key != "|": lookup[key] = (row_idx, row_dict)
    wb.close(); return lookup

def _add_team_row(row_data: dict) -> int:
    if not TEAM_EXCEL_PATH or not os.path.exists(TEAM_EXCEL_PATH):
        raise FileNotFoundError(f"Team Excel not found: {TEAM_EXCEL_PATH}")
    wb = _open_wb(TEAM_EXCEL_PATH)
    if "Database" not in wb.sheetnames: wb.create_sheet("Database")
    ws = wb["Database"]; headers = _get_headers(ws)
    if not any(headers):
        for col, h in enumerate(DB_COLUMNS, 1): ws.cell(row=1, column=col, value=h)
        headers = DB_COLUMNS
    next_row = max(ws.max_row + 1, 3)
    if "No" in headers: ws.cell(row=next_row, column=headers.index("No") + 1, value=next_row - 2)
    for col_name in SYNC_COLUMNS:
        value = row_data.get(col_name)
        if col_name in headers and value is not None:
            ws.cell(row=next_row, column=headers.index(col_name) + 1, value=value)
    wb.save(TEAM_EXCEL_PATH); wb.close(); return next_row

def _update_team_row(row_number: int, changed_fields: dict):
    wb = _open_wb(TEAM_EXCEL_PATH); ws = wb["Database"]; headers = _get_headers(ws)
    for field, vals in changed_fields.items():
        if field in headers:
            ws.cell(row=row_number, column=headers.index(field) + 1, value=vals["mine"])
    wb.save(TEAM_EXCEL_PATH); wb.close()

def _get_changed_fields(my_row: dict, team_row: dict) -> dict:
    return {col: {"mine": str(my_row.get(col) or "").strip(), "theirs": str(team_row.get(col) or "").strip()}
            for col in SYNC_COLUMNS
            if str(my_row.get(col) or "").strip() != str(team_row.get(col) or "").strip()}

def _validate_excel(path: str) -> bool:
    try:
        wb = _open_wb(path, read_only=True); wb.close(); return True
    except Exception: return False

def _get_scheduled_time() -> tuple[int, int] | None:
    return SYNC_SCHEDULE.get((TA_INCHARGE or "").lower().strip())

def _run_sync() -> dict:
    today = datetime.now().strftime("%Y-%m-%d")
    result = {"date": today, "timestamp": datetime.now().isoformat(),
              "added": 0, "updated": 0, "skipped": 0, "conflicts": 0, "errors": []}
    if not TEAM_EXCEL_PATH or not os.path.exists(TEAM_EXCEL_PATH):
        result["errors"].append(f"TEAM_EXCEL_PATH not found: {TEAM_EXCEL_PATH}"); return result
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        result["errors"].append(f"Personal EXCEL_PATH not found: {EXCEL_PATH}"); return result
    if not _validate_excel(TEAM_EXCEL_PATH):
        result["errors"].append(f"Team Excel is not a valid workbook: {TEAM_EXCEL_PATH}"); return result
    if not _validate_excel(EXCEL_PATH):
        result["errors"].append(f"Personal Excel is not a valid workbook: {EXCEL_PATH}"); return result
    personal_rows = _read_personal_db(); team_lookup = _read_team_db()
    snapshot = _load_sync_state().get("rows", {})
    email_lookup: dict[str, list[dict]] = {}
    for _k, (_rn, _rd) in team_lookup.items():
        _em = str(_rd.get("Email") or "").strip().lower()
        if _em: email_lookup.setdefault(_em, []).append({
            "request_code": str(_rd.get("Request code") or "").strip(),
            "team_ta": str(_rd.get("TA Incharge") or "").strip().lower(),
            "team_ta_display": str(_rd.get("TA Incharge") or ""),
        })
    to_sync = []
    for row in personal_rows:
        entry_date = str(row.get("Entry date") or "").strip()[:10]; key = _row_key(row)
        if not key or key == "|": continue
        if entry_date == today or (key in snapshot and _row_changed_vs_snapshot(row, snapshot[key])):
            to_sync.append(row)
    if not to_sync:
        _save_sync_state(personal_rows)
        summary = f"Daily sync {today}: No changes to sync today."
        _alert_push({"status": "sync_complete", "messages": [summary],
                        "timestamp": result["timestamp"], "details": result}); return result
    locked = False
    for my_row in to_sync:
        if locked: break
        key = _row_key(my_row)
        email = str(my_row.get("Email") or "").strip().lower()
        request_code = str(my_row.get("Request code") or "").strip()
        candidate_name = str(my_row.get("Candidate name") or "")
        serialized_row = {k: str(v) if v is not None else None for k, v in my_row.items()}
        try:
            if key not in team_lookup:
                cross_ta = [e for e in email_lookup.get(email, [])
                            if e["request_code"] != request_code
                            and e["team_ta"] != (TA_INCHARGE or "").lower().strip()] if email else []
                if cross_ta:
                    codes = ", ".join(e["request_code"] for e in cross_ta)
                    tas   = ", ".join(sorted(set(e["team_ta_display"] for e in cross_ta)))
                    _alert_push({"status": "sync_cross_ta_pending", "timestamp": datetime.now().isoformat(),
                        "email": email, "request_code": request_code, "candidate_name": candidate_name,
                        "team_ta": tas, "row_data": serialized_row,
                        "messages": [f"Cross-TA: {candidate_name} ({email}) applying across team - "
                                     f"{tas} already has this candidate for {codes}. "
                                     "Use resolve_sync_conflict to sync or skip."]})
                    result["conflicts"] += 1
                else:
                    _add_team_row(my_row); team_lookup[key] = (-1, my_row); result["added"] += 1
            else:
                team_row_num, team_row = team_lookup[key]; changed = _get_changed_fields(my_row, team_row)
                if not changed: result["skipped"] += 1; continue
                team_ta = str(team_row.get("TA Incharge") or "").strip().lower()
                my_ta   = (TA_INCHARGE or "").lower().strip()
                if team_ta != my_ta:
                    _alert_push({"status": "sync_conflict_pending", "timestamp": datetime.now().isoformat(),
                        "email": email, "request_code": request_code, "candidate_name": candidate_name,
                        "team_ta": str(team_row.get("TA Incharge") or ""), "row_data": serialized_row,
                        "messages": [f"Sync conflict: {candidate_name} ({email}) for {request_code} - "
                                     f"{team_row.get('TA Incharge')} already has this candidate in team DB. "
                                     "Use resolve_sync_conflict."]})
                    result["conflicts"] += 1
                else:
                    _update_team_row(team_row_num, changed); result["updated"] += 1
        except PermissionError:
            msg = "Team database is currently open by another user. Please close it and try again."
            result["errors"].append(msg)
            _alert_push({"status": "sync_error", "messages": [msg], "timestamp": datetime.now().isoformat()})
            locked = True
        except Exception as e:
            result["errors"].append(f"{candidate_name} ({email}): {e}")
    _save_sync_state(personal_rows)
    summary = (f"Daily sync {today}: {result['added']} added, {result['updated']} updated, "
               f"{result['skipped']} skipped, {result['conflicts']} pending conflicts.")
    if result["errors"]: summary += f" {len(result['errors'])} error(s)."
    _alert_push({"status": "sync_complete", "messages": [summary],
                    "timestamp": result["timestamp"], "details": result})
    return result

# ---------------------------------------------------------------------------
# LangGraph tools
# ---------------------------------------------------------------------------
@tool
def create_job_folder(folder_name: str) -> str:
    """Create a job folder with standard recruitment sub-folders.

    Args:
        folder_name: Name of the job folder (e.g. 'ZDA - Data Scientist - 26-ZDA-3117').
    """
    try:
        job_path = _safe_join(JOBS_BASE_DIR, folder_name)
    except ValueError:
        return f"Invalid folder name '{folder_name}': must not escape the jobs directory."
    if os.path.exists(job_path):
        return f"Folder '{folder_name}' already exists.\nFOLDER_PATH: {job_path}"
    os.makedirs(job_path)
    for sub in JOB_SUB_FOLDERS:
        os.makedirs(os.path.join(job_path, sub))
    return (f"Created '{folder_name}' with sub-folders: {', '.join(JOB_SUB_FOLDERS)}.\n"
            f"FOLDER_PATH: {job_path}")

@tool
def process_cv_file(file_path: str) -> str:
    """Process a CV file (PDF or DOCX) and add the candidate to the personal database.

    Args:
        file_path: Full path to the CV file.
    """
    result = _process_cv(file_path)
    return "\n".join([f"Status: {result['status']}"] + result["messages"])

_RESOLVABLE_STATUSES = ("duplicate", "cross_role_duplicate", "sync_conflict_pending", "sync_cross_ta_pending")

def _format_conflict_alert(alert: dict) -> str | None:
    """Build a fully-substituted conflict block (including CONFLICT_DATA) directly from
    the alert's own fields, so the LLM never has to construct or retype the JSON itself —
    that's what was producing literal "[request_code]" placeholders in the chat UI."""
    status = alert.get("status")
    if status in ("duplicate", "cross_role_duplicate"):
        email          = alert.get("email") or ""
        name           = alert.get("candidate_name") or "Unknown"
        request_code   = alert.get("request_code") or ""
        existing_job   = alert.get("existing_job") or ""
        new_job        = alert.get("new_job") or ""
        if status == "duplicate":
            issue = f"{name} already has a CV on file for this exact role ({request_code})."
        else:
            codes = ", ".join(alert.get("existing_roles") or [])
            issue = f"{name} is already in your pipeline for another role ({codes})."
        data = json.dumps({"email": email, "request_code": request_code, "type": status})
        return (f"⚠️ CONFLICT: {name} ({email})\n"
                f"Job: {request_code}\n"
                f"Already in: {existing_job}\n"
                f"New CV in: {new_job}\n"
                f"Issue: {issue}\n"
                f"CONFLICT_DATA:{data}")
    if status in ("sync_conflict_pending", "sync_cross_ta_pending"):
        email        = alert.get("email") or ""
        name         = alert.get("candidate_name") or "Unknown"
        request_code = alert.get("request_code") or ""
        team_ta      = alert.get("team_ta") or ""
        issue = f"{team_ta} already has this candidate in the team database for {request_code}."
        data = json.dumps({"email": email, "request_code": request_code, "type": "sync_conflict"})
        return (f"⚠️ CONFLICT: {name} ({email})\n"
                f"Job: {request_code}\n"
                f"Issue: {issue}\n"
                f"CONFLICT_DATA:{data}")
    return None

@tool
def get_alerts() -> str:
    """Get all pending alerts: CV processing results, duplicates, sync outcomes, errors."""
    if not _alerts: return "No pending alerts."
    lines = []
    for i, alert in enumerate(_alerts[-20:], 1):
        status = alert.get("status", "")
        block = _format_conflict_alert(alert) if status in _RESOLVABLE_STATUSES else None
        if block:
            lines.append(f"[{i}] {alert.get('timestamp', '')[:19]}")
            lines.append(block)
        else:
            lines.append(f"[{i}] {alert.get('timestamp', '')[:19]} | {status} | {alert.get('file', 'system')}")
            for msg in alert.get("messages", []): lines.append(f"    {msg}")
    return "\n".join(lines)

@tool
def clear_alerts() -> str:
    """Clear all pending alerts."""
    _alert_clear(); return "All alerts cleared."

@tool
def resolve_duplicate(email: str, request_code: str, action: str) -> str:
    """Resolve a duplicate or cross-role duplicate candidate detected during CV processing.

    Args:
        email: The candidate's email address.
        request_code: The job request code of the NEW CV being processed.
        action: 'keep' to keep the existing record unchanged, 'overwrite' to replace the
            existing record with the new CV data (same-role duplicates only), or 'add' to
            add the new CV as a separate new row without touching the existing record.
    """
    pending = next((a for a in _alerts if a.get("status") in ("duplicate", "cross_role_duplicate")
                    and str(a.get("cv_fields", {}).get("email") or "").lower() == email.strip().lower()
                    and str(a.get("folder_info", {}).get("request_code") or "") == request_code.strip()), None)
    if not pending: return f"No pending duplicate for '{email}' / '{request_code}'."
    name = pending["cv_fields"].get("full_name") or "Unknown"
    is_cross_role = pending.get("status") == "cross_role_duplicate"
    action = action.lower()
    if action == "keep":
        _alert_remove(pending); return f"Kept existing record for {name}. No changes made."
    if action == "overwrite":
        if is_cross_role:
            return "Cannot overwrite: the existing record is for a different role. Use 'add' instead."
        row_data = _build_row_data(_CVFields(**pending["cv_fields"]), pending["folder_info"], pending["file_path"])
        try: _overwrite_row(pending["duplicate_row"], row_data)
        except Exception as e: return f"Overwrite failed: {e}"
        _alert_remove(pending); return f"Overwrote row {pending['duplicate_row']} with updated data for {name}."
    if action == "add":
        row_data = _build_row_data(_CVFields(**pending["cv_fields"]), pending["folder_info"], pending["file_path"])
        try: new_row = _add_row(row_data)
        except Exception as e: return f"Failed to add row: {e}"
        _alert_remove(pending); return f"Added {name} as new row {new_row}."
    return "Invalid action. Use 'keep', 'overwrite', or 'add'."

@tool
def resolve_cross_role(email: str, new_request_code: str, action: str) -> str:
    """Resolve a cross-role duplicate: same candidate already in pipeline for a different role.

    Args:
        email: The candidate's email address.
        new_request_code: The new job request code being processed.
        action: 'add' to add the new row anyway, or 'skip' to not add.
    """
    pending = next((a for a in _alerts if a.get("status") == "cross_role_duplicate"
                    and str(a.get("cv_fields", {}).get("email") or "").lower() == email.strip().lower()
                    and str(a.get("folder_info", {}).get("request_code") or "") == new_request_code.strip()), None)
    if not pending: return f"No pending cross-role alert for '{email}' / '{new_request_code}'."
    name = pending["cv_fields"].get("full_name") or "Unknown"
    if action.lower() == "skip":
        _alert_remove(pending); return f"Skipped adding {name} for {new_request_code}. No changes made."
    if action.lower() == "add":
        row_data = _build_row_data(_CVFields(**pending["cv_fields"]), pending["folder_info"], pending["file_path"])
        try: new_row = _add_row(row_data)
        except Exception as e: return f"Failed to add row: {e}"
        _alert_remove(pending); return f"Added {name} for {new_request_code} at row {new_row}."
    return "Invalid action. Use 'add' or 'skip'."

@tool
def resolve_sync_conflict(email: str, request_code: str, action: str) -> str:
    """Resolve a pending sync conflict.

    Args:
        email: The candidate's email address.
        request_code: The job request code.
        action: 'skip' to leave team DB unchanged, or 'add_new' to insert as a new entry.
    """
    pending = next((a for a in _alerts if a.get("status") in ("sync_conflict_pending", "sync_cross_ta_pending")
                    and str(a.get("email") or "").lower() == email.strip().lower()
                    and str(a.get("request_code") or "") == request_code.strip()), None)
    if not pending: return f"No pending sync conflict for '{email}' / '{request_code}'."
    candidate_name = pending.get("candidate_name") or "Unknown"
    if action.lower() == "skip":
        _alert_remove(pending); return f"Skipped syncing {candidate_name} ({email}) for {request_code}. Team DB unchanged."
    if action.lower() == "add_new":
        row_data = {k: v for k, v in pending["row_data"].items() if v is not None}
        try: new_row = _add_team_row(row_data)
        except PermissionError: return "Team database is currently open by another user."
        except Exception as e: return f"Failed to add row: {e}"
        _alert_remove(pending); return f"Added {candidate_name} as new row {new_row} in team database."
    return "Invalid action. Use 'skip' or 'add_new'."

@tool
def run_sync_now() -> str:
    """Manually trigger the daily sync from personal database to team database right now."""
    result = _run_sync()
    lines = [f"Sync completed at {result['timestamp'][:19]}",
             f"  Added:     {result['added']}",
             f"  Updated:   {result['updated']}",
             f"  Skipped:   {result['skipped']} (no changes)",
             f"  Conflicts: {result['conflicts']}"]
    if result["errors"]: lines.append(f"  Errors: {'; '.join(result['errors'])}")
    return "\n".join(lines)

@tool
def get_sync_status() -> str:
    """Show when the last sync ran and when the next one is scheduled."""
    state = _load_sync_state()
    last  = (state.get("last_sync") or "Never")[:19]
    scheduled = _get_scheduled_time()
    sched_str = f"{scheduled[0]:02d}:{scheduled[1]:02d} daily" if scheduled else f"Not scheduled (TA '{TA_INCHARGE}' not in schedule)"
    return (f"TA Incharge:    {TA_INCHARGE}\n"
            f"Scheduled sync: {sched_str}\n"
            f"Last sync:      {last}\n"
            f"Rows tracked:   {len(state.get('rows', {}))}")

# ---------------------------------------------------------------------------
# LangGraph graph
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """You are a Recruitment Agent.

Skills:
1. create_job_folder    - create job folder with standard sub-folders
2. process_cv_file      - process a CV file and add candidate to personal database
3. get_alerts           - check pending alerts
4. clear_alerts         - dismiss resolved alerts
5. resolve_duplicate    - same email (same or different job code): keep, overwrite (same job code only), or add as new
6. resolve_cross_role   - same email + different job code: add or skip (legacy, prefer resolve_duplicate)
7. resolve_sync_conflict - sync blocked by different TA: add new row or skip
8. run_sync_now         - manually trigger personal -> team database sync
9. get_sync_status      - show last sync time and next scheduled sync

Rules:
- Never fill Stage, Status, Note, Reason for failure, or salary fields
- Never delete existing rows
- Missing CV fields: leave blank, alert recruiter
- Referrer: blank unless source = Referral
- Sync: only Excel data, never copy CV files

Folder creation:
- Always include the FOLDER_PATH line verbatim from create_job_folder result in your reply.

Sync workflow:
- After run_sync_now() completes, ALWAYS call get_alerts() immediately as your next step.
- Report in plain language: "Sync complete: X added, Y updated, Z unchanged, N conflicts need attention."
- If conflicts exist, show them immediately — never ask "would you like to check alerts?".
- get_alerts() already formats each conflict in full, including the CONFLICT_DATA line with
  the real email/request_code/type values filled in. Copy every conflict block from
  get_alerts() into your reply EXACTLY as returned, character for character — including the
  CONFLICT_DATA line. Never retype, reformat, summarize, or recompute the CONFLICT_DATA line
  yourself; never write a placeholder like "[request_code]" — always use get_alerts()'s own
  text verbatim.
- After listing all conflicts, say: "Use the buttons above to resolve each conflict."

Proactive suggestions (add 1–2 sentences after completing each task):
- After create_job_folder → suggest uploading CVs to the new folder.
- After process_cv_file (success) → suggest running a sync to push to the team database.
- After run_sync_now → suggest checking alerts if any conflicts were flagged.
- After resolving a conflict → suggest viewing the Excel file to confirm the change.
Keep suggestions brief (1 sentence each), natural, and only when they add value. Do not repeat the same suggestion twice in a row."""


class State(TypedDict):
    messages: Annotated[list, add_messages]


tools = [
    create_job_folder,
    process_cv_file, get_alerts, clear_alerts,
    resolve_duplicate, resolve_cross_role, resolve_sync_conflict,
    run_sync_now, get_sync_status,
]
llm_with_tools = llm.bind_tools(tools)


def chatbot(state: State) -> dict:
    messages = [SystemMessage(content=SYSTEM_PROMPT)] + state["messages"]
    return {"messages": [llm_with_tools.invoke(messages)]}


graph_builder = StateGraph(State)
graph_builder.add_node("chatbot", chatbot)
graph_builder.add_node("tools", ToolNode(tools))
graph_builder.add_edge(START, "chatbot")
graph_builder.add_conditional_edges("chatbot", tools_condition)
graph_builder.add_edge("tools", "chatbot")
graph = graph_builder.compile()

# ---------------------------------------------------------------------------
# HTTP handlers
# ---------------------------------------------------------------------------
_ALLOWED_EXT = {".pdf", ".docx"}

# job_id -> {"status", "total", "progress", "done", "failed", "current", "results", ...}
_parse_jobs: dict[str, dict] = {}

_LINKIFY_RE = re.compile(
    r'FOLDER_PATH:\s*([A-Za-z]:\\[^\n\r]+)'   # group 1 — folder path with explicit prefix
    r'|'
    r'([A-Za-z]:\\(?:[^\n\r"\'<>|*?]+))'      # group 2 — bare Windows path
)

def linkify_paths(text: str) -> str:
    """Wrap Windows paths in markers so the UI can render them as links.

    FOLDER_PATH: prefix  → __FOLDER__...__ENDFOLDER__  (Open Folder button)
    Bare path            → __LINK__...__URL__...__ENDLINK__  (copy button)
    """
    def _replace(m: re.Match) -> str:
        if m.group(1) is not None:
            raw = m.group(1).rstrip(".,;:)'\"")
            return f"__FOLDER__{raw}__ENDFOLDER__"
        raw = m.group(2).rstrip(".,;:)'\"")
        url = "file:///" + raw.replace("\\", "/")
        return f"__LINK__{raw}__URL__{url}__ENDLINK__"
    return _LINKIFY_RE.sub(_replace, text)


_FOLDER_PATH_LINE_RE   = re.compile(r'FOLDER_PATH:\s*[A-Za-z]:\\[^\n\r]+')
_CONFLICT_DATA_LINE_RE = re.compile(r'CONFLICT_DATA:\{[^\n\r]+\}')

def _ensure_tool_markers(messages: list, final_text: str) -> str:
    """Guarantee FOLDER_PATH / CONFLICT_DATA lines produced by tool calls (create_job_folder,
    get_alerts) survive into the final reply, even if the LLM paraphrased or dropped them
    when writing its prose response. Without this, the chat UI's Open Folder button and
    conflict-resolution buttons silently fail to render."""
    extra_lines: list[str] = []
    for m in messages:
        if not isinstance(m, ToolMessage):
            continue
        content = m.content if isinstance(m.content, str) else str(m.content)
        for line_re in (_FOLDER_PATH_LINE_RE, _CONFLICT_DATA_LINE_RE):
            for match in line_re.finditer(content):
                line = match.group(0)
                if line not in final_text and line not in extra_lines:
                    extra_lines.append(line)
    if not extra_lines:
        return final_text
    return final_text.rstrip() + "\n" + "\n".join(extra_lines)


async def handle_list_folders(request: Request) -> JSONResponse:
    base = Path(JOBS_BASE_DIR)
    folders: list[str] = sorted(d.name for d in base.iterdir() if d.is_dir()) if base.is_dir() else []
    return JSONResponse({"folders": folders})


async def handle_open_folder(request: Request) -> JSONResponse:
    path = request.query_params.get("path", "").strip()
    if not path:
        return JSONResponse({"status": "error", "response": "Missing path"}, status_code=400)
    try:
        path_obj = _safe_join(JOBS_BASE_DIR, os.path.normpath(path))
    except ValueError:
        return JSONResponse({"status": "error", "response": "Path outside allowed directory"}, status_code=403)
    if not path_obj.is_dir():
        return JSONResponse({"status": "error", "response": f"Not a directory: {path}"}, status_code=404)
    try:
        subprocess.Popen(["explorer", str(path_obj)])
        return JSONResponse({"status": "ok"})
    except Exception as e:
        return JSONResponse({"status": "error", "response": str(e)}, status_code=500)


async def handle_invocations(request: Request) -> JSONResponse:
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"status": "error", "response": "Request body must be valid JSON."}, status_code=400)

    message = str(body.get("message", "")).strip()
    if not message:
        return JSONResponse({"status": "error", "response": "Field 'message' is required."}, status_code=400)

    try:
        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(
            None, lambda: graph.invoke({"messages": [("user", message)]})
        )
        final_text = _ensure_tool_markers(result["messages"], result["messages"][-1].content)
        return JSONResponse({
            "status": "success",
            "response": linkify_paths(final_text),
            "timestamp": datetime.now().isoformat(),
        })
    except Exception as exc:
        return JSONResponse(
            {"status": "error", "response": f"Agent error: {exc}", "timestamp": datetime.now().isoformat()},
            status_code=500,
        )


async def handle_health(request: Request) -> JSONResponse:
    return JSONResponse({"status": "healthy", "jobs_dir": JOBS_BASE_DIR,
                         "timestamp": datetime.now().isoformat()})


# ---------------------------------------------------------------------------
# Phase-1 / Phase-2 handlers (fast-save + background parse)
# ---------------------------------------------------------------------------

def _parse_one(path: str, timeout: float = 120.0) -> dict:
    """Run _process_cv in a sub-thread so we can apply a wall-clock timeout."""
    result: dict = {}
    exc_holder: list = []

    def _run() -> None:
        try:
            result.update(_process_cv(path))
        except Exception as e:  # noqa: BLE001
            exc_holder.append(e)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    t.join(timeout=timeout)
    if t.is_alive():
        return {"status": "timeout", "messages": [f"Parsing timed out after {int(timeout)}s."]}
    if exc_holder:
        raise exc_holder[0]
    return result or {"status": "error", "messages": ["No result returned."]}


def _parse_worker(job_id: str, file_paths: list[str]) -> None:
    """Background thread: parse up to 3 CVs in parallel, then flush Excel once."""
    job = _parse_jobs[job_id]

    def _run_one(path: str) -> dict:
        name = os.path.basename(path)
        job["current"] = name
        try:
            result = _parse_one(path)
            return {"name": name, **result, "path": path}
        except Exception as e:  # noqa: BLE001
            return {"name": name, "status": "error", "messages": [str(e)], "path": path}

    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {pool.submit(_run_one, p): p for p in file_paths}
        for fut in as_completed(futures):
            result = fut.result()
            name   = result.get("name", os.path.basename(futures[fut]))
            status = result.get("status", "error")
            messages = result.get("messages", [])
            if status in ("success", "no_excel", "duplicate", "cross_role_duplicate"):
                job["done"] += 1
            else:
                job["failed"] += 1
            job["results"].append({**result, "name": name, "status": status, "messages": messages})
            job["progress"] += 1

    _flush_excel_pending()
    job["status"] = "complete" if job["failed"] == 0 else ("partial" if job["done"] > 0 else "failed")
    job["finished_at"] = datetime.now().isoformat()


async def handle_save_cv(request: Request) -> JSONResponse:
    """Phase 1 — save a CV to disk immediately; no AI parsing."""
    try:
        form = await request.form()
    except Exception as exc:
        return JSONResponse({"status": "error", "response": f"Could not parse form: {exc}"}, status_code=400)

    folder    = str(form.get("folder")    or "").strip()
    subfolder = str(form.get("subfolder") or "").strip()
    upload    = form.get("file")

    if not folder or not subfolder or upload is None:
        return JSONResponse(
            {"status": "error", "response": "Fields 'folder', 'subfolder', and 'file' are required."},
            status_code=400,
        )
    if subfolder not in JOB_SUB_FOLDERS:
        return JSONResponse(
            {"status": "error", "response": f"'subfolder' must be one of: {', '.join(JOB_SUB_FOLDERS)}"},
            status_code=400,
        )

    raw_name = Path(upload.filename).name
    safe_name = unicodedata.normalize("NFC", raw_name)
    ext = Path(safe_name).suffix.lower()
    if ext not in _ALLOWED_EXT:
        return JSONResponse(
            {"status": "error", "response": f"File type '{ext}' not supported. Use PDF or DOCX."},
            status_code=400,
        )

    try:
        save_dir = _safe_join(JOBS_BASE_DIR, folder, subfolder)
    except ValueError:
        return JSONResponse({"status": "error", "response": "Invalid folder path."}, status_code=400)
    save_dir.mkdir(parents=True, exist_ok=True)
    save_path = save_dir / safe_name

    contents = await upload.read()
    with open(save_path, "wb") as fh:
        fh.write(contents)

    return JSONResponse({
        "status":    "saved",
        "name":      safe_name,
        "saved_to":  str(save_path),
        "folder":    folder,
        "subfolder": subfolder,
        "timestamp": datetime.now().isoformat(),
    })


async def handle_start_parse(request: Request) -> JSONResponse:
    """Phase 2 — kick off background parsing for a list of already-saved file paths."""
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"status": "error", "response": "Request body must be valid JSON."}, status_code=400)

    files: list[str] = body.get("files", [])
    if not files:
        return JSONResponse({"status": "error", "response": "No files to parse."}, status_code=400)

    job_id = uuid.uuid4().hex[:12]
    _parse_jobs[job_id] = {
        "status":      "running",
        "total":       len(files),
        "progress":    0,
        "done":        0,
        "failed":      0,
        "current":     "",
        "results":     [],
        "started_at":  datetime.now().isoformat(),
        "finished_at": None,
    }

    threading.Thread(target=_parse_worker, args=(job_id, files), daemon=True).start()
    return JSONResponse({"job_id": job_id, "total": len(files), "status": "running"})


async def handle_parse_status(request: Request) -> JSONResponse:
    """Poll the current status of a background parse job."""
    job_id = request.path_params.get("job_id", "")
    job = _parse_jobs.get(job_id)
    if not job:
        return JSONResponse({"status": "error", "response": f"Unknown job_id: {job_id!r}"}, status_code=404)
    return JSONResponse({"job_id": job_id, **job})


# Load Excel rows into memory so duplicate checks are O(1) during CV parsing
_init_excel_cache()

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = Starlette(
    routes=[
        Route("/invocations",           handle_invocations,   methods=["POST"]),
        Route("/save-cv",               handle_save_cv,       methods=["POST"]),
        Route("/start-parse",           handle_start_parse,   methods=["POST"]),
        Route("/parse-status/{job_id}", handle_parse_status,  methods=["GET"]),
        Route("/list-folders",          handle_list_folders,  methods=["GET"]),
        Route("/open-folder",           handle_open_folder,   methods=["GET"]),
        Route("/health",                handle_health,        methods=["GET"]),
    ],
    middleware=[
        Middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["GET", "POST", "OPTIONS"], allow_headers=["*"])
    ],
)

if __name__ == "__main__":
    print()
    print("  Recruiter Agent - Local Server")
    print(f"  Jobs dir : {JOBS_BASE_DIR}")
    print(f"  URL      : http://0.0.0.0:8080")
    print(f"  Tunnel   : ngrok http 8080")
    print()
    uvicorn.run(app, host="0.0.0.0", port=8080, log_level="info")
